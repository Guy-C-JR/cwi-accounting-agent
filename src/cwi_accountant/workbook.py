from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
import shutil
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.worksheet import Worksheet

from cwi_accountant.models import ProposedExpenseEntry, RecurringBillCandidate, VendorCandidate
from cwi_accountant.services.category_mapper import CanonicalLists
from cwi_accountant.utils import normalize_vendor_name


@dataclass(slots=True)
class WriteResult:
    sheet_name: str
    row_number: int
    expense_ref: str | None = None


class WorkbookGateway:
    def __init__(self, workbook_path: Path, backups_dir: Path):
        self.workbook_path = workbook_path
        self.backups_dir = backups_dir
        self.backups_dir.mkdir(parents=True, exist_ok=True)

    def backup_workbook(self, reason: str) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target = self.backups_dir / f"{self.workbook_path.stem}.{reason}.{timestamp}.xlsx"
        shutil.copy2(self.workbook_path, target)
        return target

    def load_lists(self) -> CanonicalLists:
        wb = load_workbook(self.workbook_path)
        ws = wb["Lists"]

        def column_values(col: str) -> list[str]:
            values: list[str] = []
            for cell in ws[col][1:]:
                if cell.value in (None, ""):
                    continue
                values.append(str(cell.value).strip())
            return values

        return CanonicalLists(
            categories=column_values("A"),
            subcategories=column_values("B"),
            payment_methods=column_values("C"),
            yes_no=column_values("D"),
            vendor_types=column_values("E"),
            frequencies=column_values("F"),
            statuses=column_values("H"),
        )

    def find_expense_duplicate(
        self,
        *,
        vendor: str | None,
        date_value: Any,
        amount: Decimal | None,
    ) -> int | None:
        wb = load_workbook(self.workbook_path, data_only=True)
        ws = wb["Expense_Log"]
        if not vendor or date_value is None or amount is None:
            return None

        target_vendor = normalize_vendor_name(vendor)
        for row in range(2, ws.max_row + 1):
            row_vendor = normalize_vendor_name(str(ws.cell(row=row, column=4).value or ""))
            row_date = ws.cell(row=row, column=1).value
            row_amount = ws.cell(row=row, column=10).value
            if row_vendor != target_vendor:
                continue
            if row_date and hasattr(row_date, "date"):
                row_date = row_date.date()
            if str(row_date) != str(date_value):
                continue
            if row_amount is None:
                continue
            if Decimal(str(row_amount)).quantize(Decimal("0.01")) == amount:
                return row
        return None

    def append_expense(self, entry: ProposedExpenseEntry) -> WriteResult:
        self.backup_workbook("expense_write")
        wb = load_workbook(self.workbook_path)
        ws = wb["Expense_Log"]

        headers = self._header_map(ws)
        target_row = ws.max_row + 1
        template_row = max(2, target_row - 1)

        self._copy_row_pattern(ws, template_row, target_row, ws.max_column)

        direct_map = {
            "Date": entry.date,
            "Vendor": entry.vendor,
            "Category": entry.category,
            "Subcategory": entry.subcategory,
            "Description": entry.description,
            "Payment Method": entry.payment_method,
            "Account/Card": entry.account_card,
            "Amount ($)": float(entry.amount) if entry.amount is not None else None,
            "Tax Deductible?": entry.tax_deductible,
            "Receipt?": entry.receipt,
            "Receipt Link/File": entry.receipt_link_file,
            "Business Purpose": entry.business_purpose,
            "Billable to Client?": entry.billable_to_client,
            "Client/Project": entry.client_project,
            "Recurring?": entry.recurring,
            "Notes": entry.notes,
        }

        for header, value in direct_map.items():
            col = headers.get(header)
            if not col:
                continue
            self._write_if_not_formula(ws, target_row, col, value)

        wb.save(self.workbook_path)
        return WriteResult(sheet_name="Expense_Log", row_number=target_row, expense_ref=f"Expense_Log!{target_row}")

    def upsert_vendor(self, vendor: VendorCandidate) -> WriteResult:
        self.backup_workbook("vendor_write")
        wb = load_workbook(self.workbook_path)
        ws = wb["Vendors"]
        headers = self._header_map(ws)

        existing_row = self._find_vendor_row(ws, vendor.vendor_name)
        if existing_row is None:
            target_row = ws.max_row + 1
            template_row = max(2, target_row - 1)
            self._copy_row_pattern(ws, template_row, target_row, ws.max_column)
        else:
            target_row = existing_row

        mapping = {
            "Vendor Name": vendor.vendor_name,
            "Vendor Type": vendor.vendor_type,
            "Contact Person": vendor.contact_person,
            "Email": vendor.email,
            "Phone": vendor.phone,
            "Address": vendor.address,
            "Website": vendor.website,
            "Tax Form Needed?": vendor.tax_form_needed,
            "1099 Eligible?": vendor.eligible_1099,
            "Usual Category": vendor.usual_category,
            "Payment Terms": vendor.payment_terms,
            "Status": vendor.status,
            "Notes": vendor.notes,
        }

        for header, value in mapping.items():
            col = headers.get(header)
            if not col:
                continue
            cell = ws.cell(row=target_row, column=col)
            if existing_row is not None and cell.value not in (None, "") and value in (None, ""):
                continue
            self._write_if_not_formula(ws, target_row, col, value)

        wb.save(self.workbook_path)
        return WriteResult(sheet_name="Vendors", row_number=target_row)

    def upsert_recurring_bill(self, recurring: RecurringBillCandidate) -> WriteResult:
        self.backup_workbook("recurring_write")
        wb = load_workbook(self.workbook_path)
        ws = wb["Recurring_Bills"]
        headers = self._header_map(ws)

        existing_row = self._find_recurring_row(ws, recurring.vendor, recurring.expense_name)
        if existing_row is None:
            target_row = ws.max_row + 1
            template_row = max(2, target_row - 1)
            self._copy_row_pattern(ws, template_row, target_row, ws.max_column)
        else:
            target_row = existing_row

        annualized = None
        if recurring.amount is not None and recurring.frequency:
            freq = recurring.frequency.lower()
            if "month" in freq:
                annualized = float(recurring.amount * Decimal("12"))
            elif "quarter" in freq:
                annualized = float(recurring.amount * Decimal("4"))
            elif "annual" in freq or "year" in freq:
                annualized = float(recurring.amount)

        mapping = {
            "Vendor": recurring.vendor,
            "Expense Name": recurring.expense_name,
            "Category": recurring.category,
            "Amount ($)": float(recurring.amount) if recurring.amount is not None else None,
            "Frequency": recurring.frequency,
            "Due Day": recurring.due_day,
            "Start Date": recurring.first_seen,
            "End Date": None,
            "Active?": "Yes",
            "Annualized Cost ($)": annualized,
            "Notes": f"Candidate confidence: {recurring.confidence:.2f}",
        }

        for header, value in mapping.items():
            col = headers.get(header)
            if col:
                self._write_if_not_formula(ws, target_row, col, value)

        wb.save(self.workbook_path)
        return WriteResult(sheet_name="Recurring_Bills", row_number=target_row)

    def append_receipt_index(
        self,
        *,
        receipt_id: str,
        date_value: Any,
        vendor: str | None,
        amount: Decimal | None,
        linked_expense_ref: str | None,
        file_path: str,
        verified: str,
        notes: str,
    ) -> WriteResult:
        self.backup_workbook("receipt_write")
        wb = load_workbook(self.workbook_path)
        ws = wb["Receipt_Index"]
        headers = self._header_map(ws)

        target_row = ws.max_row + 1
        template_row = max(2, target_row - 1)
        self._copy_row_pattern(ws, template_row, target_row, ws.max_column)

        mapping = {
            "Receipt ID": receipt_id,
            "Date": date_value,
            "Vendor": vendor,
            "Amount ($)": float(amount) if amount is not None else None,
            "Linked Expense Ref": linked_expense_ref,
            "File Name / Link": file_path,
            "Stored Where": str(Path(file_path).parent),
            "Verified?": verified,
            "Notes": notes,
        }
        for header, value in mapping.items():
            col = headers.get(header)
            if col:
                self._write_if_not_formula(ws, target_row, col, value)

        wb.save(self.workbook_path)
        return WriteResult(sheet_name="Receipt_Index", row_number=target_row)

    def get_expense_row_snapshot(self, row: int) -> dict[str, Any]:
        wb = load_workbook(self.workbook_path, data_only=False)
        ws = wb["Expense_Log"]
        headers = self._header_map(ws)
        snapshot: dict[str, Any] = {}
        for header, col in headers.items():
            snapshot[header] = ws.cell(row=row, column=col).value
        return snapshot

    def _find_vendor_row(self, ws: Worksheet, vendor_name: str) -> int | None:
        target = normalize_vendor_name(vendor_name)
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if normalize_vendor_name(str(name or "")) == target:
                return row
        return None

    def _find_recurring_row(self, ws: Worksheet, vendor: str, expense_name: str) -> int | None:
        target_vendor = normalize_vendor_name(vendor)
        target_name = (expense_name or "").strip().lower()
        for row in range(2, ws.max_row + 1):
            row_vendor = normalize_vendor_name(str(ws.cell(row=row, column=1).value or ""))
            row_name = str(ws.cell(row=row, column=2).value or "").strip().lower()
            if row_vendor == target_vendor and row_name == target_name:
                return row
        return None

    def _copy_row_pattern(self, ws: Worksheet, from_row: int, to_row: int, max_col: int) -> None:
        for col in range(1, max_col + 1):
            source = ws.cell(row=from_row, column=col)
            target = ws.cell(row=to_row, column=col)
            if source.has_style:
                target._style = copy(source._style)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
            target.alignment = copy(source.alignment)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)

            if isinstance(source.value, str) and source.value.startswith("="):
                try:
                    target.value = Translator(source.value, origin=source.coordinate).translate_formula(
                        target.coordinate
                    )
                except Exception:
                    target.value = source.value

    def _write_if_not_formula(self, ws: Worksheet, row: int, col: int, value: Any) -> None:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            return
        cell.value = value

    @staticmethod
    def _header_map(ws: Worksheet) -> dict[str, int]:
        out: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value in (None, ""):
                continue
            out[str(value).strip()] = col
        return out
