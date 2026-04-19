from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from cwi_accountant.models import ProposedExpenseEntry
from cwi_accountant.workbook import WorkbookGateway


def _build_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense_Log"
    headers = [
        "Date", "Month", "Year", "Vendor", "Category", "Subcategory", "Description", "Payment Method",
        "Account/Card", "Amount ($)", "Tax Deductible?", "Receipt?", "Receipt Link/File", "Business Purpose",
        "Billable to Client?", "Client/Project", "Recurring?", "Notes", "Quarter", "Week #", "Month Key"
    ]
    ws.append(headers)
    ws.append([
        date(2026, 1, 1),
        "=TEXT(A2,\"mmm\")",
        "=YEAR(A2)",
        "Existing Vendor",
        "Software / SaaS",
        "AI/API Usage",
        "Seed row",
        "Credit Card",
        "Amex",
        10.0,
        "Yes",
        "Yes",
        "/tmp/seed.pdf",
        "Ops",
        "No",
        "",
        "No",
        "seed",
        "=\"Q\"&ROUNDUP(MONTH(A2)/3,0)",
        "=ISOWEEKNUM(A2)",
        "=TEXT(A2,\"yyyy-mm\")",
    ])

    for name, row in [
        ("Vendors", ["Vendor Name", "Vendor Type", "Contact Person", "Email", "Phone", "Address", "Website", "Tax Form Needed?", "1099 Eligible?", "Usual Category", "Payment Terms", "Status", "Notes"]),
        ("Recurring_Bills", ["Vendor", "Expense Name", "Category", "Amount ($)", "Frequency", "Due Day", "Payment Method", "Auto-Pay?", "Start Date", "End Date", "Active?", "Annualized Cost ($)", "Notes"]),
        ("Receipt_Index", ["Receipt ID", "Date", "Vendor", "Amount ($)", "Linked Expense Ref", "File Name / Link", "Stored Where", "Verified?", "Notes"]),
        ("Lists", ["Categories", "Subcategories", "Payment_Methods", "Yes_No", "Vendor_Types", "Frequency", "Tax_Treatment", "Status", "Months"]),
    ]:
        sheet = wb.create_sheet(name)
        sheet.append(row)
    lists = wb["Lists"]
    lists.append(["Software / SaaS", "AI/API Usage", "Credit Card", "Yes", "Professional Service", "Monthly", "Yes", "Active", "Jan"])
    lists.append(["Advertising / Marketing", "Subscription", "ACH", "No", "Other", "Annual", "No", "Inactive", "Feb"])

    wb.save(path)


def test_append_expense_and_backup(tmp_path: Path) -> None:
    workbook_path = tmp_path / "tracker.xlsx"
    backups = tmp_path / "backups"
    _build_workbook(workbook_path)

    gateway = WorkbookGateway(workbook_path, backups)
    entry = ProposedExpenseEntry(
        date=date(2026, 2, 1),
        vendor="OpenAI",
        category="Software / SaaS",
        subcategory="AI/API Usage",
        description="API usage",
        payment_method="Credit Card",
        account_card="Amex",
        amount="42.50",
        tax_deductible="Yes",
        receipt="Yes",
        receipt_link_file="/tmp/openai.pdf",
    )

    result = gateway.append_expense(entry)
    assert result.row_number == 3
    assert any(backups.glob("*.xlsx"))

    wb = load_workbook(workbook_path)
    ws = wb["Expense_Log"]
    assert ws["D3"].value == "OpenAI"
    assert ws["J3"].value == 42.5
    assert isinstance(ws["B3"].value, str) and ws["B3"].value.startswith("=")
